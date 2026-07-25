#!/usr/bin/env python3
"""persist_results.py — write screen results into all_results.xlsx (reimplementation rows).

Usage: persist_results.py <RUN_CSV> [ALL_RESULTS_XLSX]
For each (method,agents,freq) in RUN_CSV it updates the matching source=reimplementation row:
  status ok  -> makespan/swt/runtime_s + recomputed ms_gap%/swt_gap%/rt_gap% vs mapped original
  otherwise  -> all N/A
Gaps are stored as DECIMAL FRACTIONS (e.g. 0.026 for +2.6%) and the cells are given a
percent number format, matching how all_results.xlsx displays them.
Gaps use the same per-step-ms runtime conversion as compare_method.py for *-MLA* references.
Backs up ALL_RESULTS_XLSX to .bak_persist before writing. Leaves untouched any cell not in RUN_CSV.
"""
import csv, sys, shutil
import openpyxl

DEFAULT_ALL = "/Users/jiaqit/Desktop/meta/paper/MAPD_framework_imp/all_results.xlsx"

COMPARE = {
    "TP-STA*": "TP-STA*",
    "TPTS-STA*": "TPTS-STA*",
    "CENTRAL-ECBS": "CENTRAL-ECBS",
    "CENTRAL-ECBS-SIPP": "CENTRAL-ECBS",
    "TA-Hybrid-STA*": "TA-Hybrid-STA*",
    "TA-Prioritized-STA*": "TA-Prioritized-STA*",
    "Hungarian+PBS-MLA*": "Hungarian+PBS-MLA*",
    "Hungarian+wPBS-MLA*": "Hungarian+wPBS-MLA*",
    "LNS(1s)+PBS-MLA*": "LNS(1s)+PBS-MLA*",
    "LNS(1s)+wPBS-MLA*": "LNS(1s)+wPBS-MLA*",
    "Hungarian+PBS-MLSIPP": "Hungarian+PBS-MLA*",
    "Hungarian+wPBS-MLSIPP": "Hungarian+wPBS-MLA*",
    "LNS(1s)+PBS-MLSIPP": "LNS(1s)+PBS-MLA*",
    "LNS(1s)+wPBS-MLSIPP": "LNS(1s)+wPBS-MLA*",
    "TP-SIPP": "TP-STA*",
    "TPTS-SIPP": "TPTS-STA*",
    "Hungarian+PP-SIPP": "Hungarian+PBS-MLA*",
    "LNS(1s)+PP-SIPP": "LNS(1s)+PBS-MLA*",
}
PER_STEP = {
    "Hungarian+PBS-MLA*",
    "Hungarian+wPBS-MLA*",
    "LNS(1s)+PBS-MLA*",
    "LNS(1s)+wPBS-MLA*",
}
PCT_FMT = "0.00%"


def num(x):
    try:
        return float(x)
    except Exception:
        return None


def norm_key(x):
    """Canonicalize an agents/freq key so xlsx numbers (10, 0.2, 2.0) and csv
    strings ('10','0.2','2') compare equal."""
    s = str(x).strip()
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else ("%g" % f)
    except Exception:
        return s


def main():
    if len(sys.argv) < 2:
        sys.exit("usage: persist_results.py <RUN_CSV> [ALL_RESULTS_XLSX]")
    run_csv = sys.argv[1]
    allp = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_ALL
    shutil.copy(allp, allp + ".bak_persist")

    wb = openpyxl.load_workbook(allp)  # keep formatting; do NOT use data_only for saving
    ws = wb.active
    hdr = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(hdr)}  # 0-based
    need = ("method", "agents", "freq", "makespan", "swt", "runtime_s",
            "ms_gap%", "swt_gap%", "rt_gap%", "source")
    missing = [k for k in need if k not in col]
    if missing:
        sys.exit(f"ERROR: {allp} missing columns {missing}; found {hdr}")

    def cell(r, name):  # r is a tuple of cell objects (1-based worksheet row)
        return r[col[name]]

    # index originals for gap computation
    orig = {}
    for row in ws.iter_rows(min_row=2):
        if str(cell(row, "source").value) != "original":
            continue
        k = (str(cell(row, "method").value).strip(),
             norm_key(cell(row, "agents").value),
             norm_key(cell(row, "freq").value))
        orig[k] = (num(cell(row, "makespan").value),
                   num(cell(row, "swt").value),
                   num(cell(row, "runtime_s").value))

    # load fresh run results (RUN_CSV: method,agents,freq,makespan,swt,runtime_s,status)
    fresh = {}
    rc = list(csv.reader(open(run_csv)))
    if not rc:
        sys.exit(f"ERROR: {run_csv} is empty")
    rhdr = [h.strip() for h in rc[0]]
    ri = {h: i for i, h in enumerate(rhdr)}
    pos = all(h in ri for h in ("method", "agents", "freq", "makespan", "swt", "runtime_s"))
    for r in rc[1:]:
        if len(r) < 6:
            continue
        if pos:
            m, ag, fr = r[ri["method"]], r[ri["agents"]], r[ri["freq"]]
            ms, swt, rt = r[ri["makespan"]], r[ri["swt"]], r[ri["runtime_s"]]
            st = r[ri["status"]] if "status" in ri and len(r) > ri["status"] else ""
        else:  # positional fallback
            m, ag, fr, ms, swt, rt = r[0], r[1], r[2], r[3], r[4], r[5]
            st = r[6] if len(r) > 6 else ""
        fresh[(m.strip(), norm_key(ag), norm_key(fr))] = (ms, swt, rt, st)

    def gap_fractions(m, ms, swt, rt):
        """Return (ms_gap, swt_gap, rt_gap) as signed decimal fractions, or Nones."""
        og = COMPARE.get(m)
        ov = orig.get((og, key[1], key[2])) if og else None
        if ov is None or None in ov:
            return (None, None, None)
        omS, oS, oR = ov
        oR_s = oR * omS / 1000.0 if og in PER_STEP else oR

        def g(a, b):
            a = num(a)
            return None if a is None or b in (None, 0) else (a - b) / b

        return (g(ms, omS), g(swt, oS), g(rt, oR_s))

    upd = na = 0
    for row in ws.iter_rows(min_row=2):
        if str(cell(row, "source").value) != "reimplementation":
            continue
        key = (str(cell(row, "method").value).strip(),
               norm_key(cell(row, "agents").value),
               norm_key(cell(row, "freq").value))
        if key not in fresh:
            continue
        ms, swt, rt, st = fresh[key]
        gap_cells = [cell(row, "ms_gap%"), cell(row, "swt_gap%"), cell(row, "rt_gap%")]
        if st in ("ok", "") and num(ms) is not None:
            cell(row, "makespan").value = num(ms)
            cell(row, "swt").value = num(swt)
            cell(row, "runtime_s").value = num(rt)
            for c, gv in zip(gap_cells, gap_fractions(key[0], ms, swt, rt)):
                if gv is None:
                    c.value = "-"
                else:
                    c.value = gv
                    c.number_format = PCT_FMT
            upd += 1
        else:
            cell(row, "makespan").value = "N/A"
            cell(row, "swt").value = "N/A"
            cell(row, "runtime_s").value = "N/A"
            for c in gap_cells:
                c.value = "N/A"
            na += 1

    wb.save(allp)
    print(f"persisted {upd} cells ({na} N/A) from {run_csv} -> {allp}")


if __name__ == "__main__":
    main()
